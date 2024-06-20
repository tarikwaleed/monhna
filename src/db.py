from flask_sqlalchemy import SQLAlchemy
from app import app

import os
from app import app
from openpyxl import load_workbook
from datetime import datetime

from dotenv import find_dotenv, load_dotenv

load_dotenv(find_dotenv(), override=True)
DATA_PATH = os.getenv("DATA_PATH")


db = SQLAlchemy(app)


class Coupon(db.Model):
    __tablename__ = "coupons"

    id = db.Column(db.Integer, primary_key=True)
    request_id = db.Column(db.String)
    prefix = db.Column(db.String)
    coupon_cde=db.Column(db.String)
    advertiser = db.Column(db.String)
    num_coupons = db.Column(db.Integer)
    status = db.Column(db.String)
    created_at = db.Column(db.DateTime)
    user_id = db.Column(db.String, default="")


def populate_database():
    wb = load_workbook(DATA_PATH)
    sheet = wb.active

    with app.app_context():
        for row in sheet.iter_rows(min_row=2, values_only=True):
            (
                request_id,
                prefix,
                coupon_code,
                advertiser,
                num_coupons,
                status,
                created_at,
                _,
            ) = row
            coupon = Coupon(
                request_id=request_id,
                prefix=prefix,
                coupon_code=coupon_code,
                advertiser=advertiser,
                num_coupons=num_coupons,
                status=status,
                created_at=datetime.strptime(created_at, "%Y-%m-%dT%H:%M:%SZ"),
                user_id=None,
            )
            db.session.add(coupon)

        db.session.commit()

    print("Database populated successfully.")


if __name__ == "__main__":
    with app.app_context():
        db.create_all()
    populate_database()
